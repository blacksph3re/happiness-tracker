import sqlite3
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from tests.conftest import make_user

HOUR = 3600


def at(day, hour, minute=0):
    """Build an instant on a day in June 2026, as the API takes it."""
    return datetime(2026, 6, day, hour, minute).isoformat()


def make_project(client, headers, name="The rewrite", **extra):
    """Create a project and return its body."""
    response = client.post(
        "/api/projects", headers=headers, json={"name": name, **extra}
    )
    assert response.status_code == 201, response.text
    return response.json()


def make_tag(client, headers, name="Work", **extra):
    """Create a tag and return its body."""
    response = client.post("/api/tags", headers=headers, json={"name": name, **extra})
    assert response.status_code == 201, response.text
    return response.json()


def check_in(client, headers, project_id, day=10, hour=9, offset=0):
    """Start a timer."""
    return client.post(
        f"/api/projects/{project_id}/check-in",
        headers=headers,
        json={"at": at(day, hour), "utc_offset": offset},
    )


def check_out(client, headers, project_id, day=10, hour=17):
    """Stop a timer."""
    return client.post(
        f"/api/projects/{project_id}/check-out",
        headers=headers,
        json={"at": at(day, hour)},
    )


def record(client, headers, project_id, start, end, offset=0):
    """Add a finished session by hand."""
    return client.post(
        "/api/time/entries",
        headers=headers,
        json={
            "project_id": project_id,
            "started_at": start,
            "ended_at": end,
            "utc_offset": offset,
        },
    )


def totals(client, headers, by="project", **params):
    """Read the summary as ``{(day, key): seconds}``."""
    response = client.get(
        "/api/time/summary", headers=headers, params={"by": by, **params}
    )
    assert response.status_code == 200, response.text
    return {(row["day"], row["key"]): row["seconds"] for row in response.json()}


def test_a_new_account_has_no_projects(client, admin_headers):
    assert client.get("/api/projects", headers=admin_headers).json() == []


def test_two_projects_run_at_once(client, admin_headers):
    work = make_project(client, admin_headers, "Work")
    meeting = make_project(client, admin_headers, "Meeting")
    assert check_in(client, admin_headers, work["id"], hour=9).status_code == 201
    assert check_in(client, admin_headers, meeting["id"], hour=11).status_code == 201

    running = [
        e
        for e in client.get("/api/time/entries", headers=admin_headers).json()
        if e["ended_at"] is None
    ]
    assert len(running) == 2


def test_checking_into_a_running_project_twice_is_refused(client, admin_headers):
    project = make_project(client, admin_headers)
    check_in(client, admin_headers, project["id"])
    assert check_in(client, admin_headers, project["id"], hour=10).status_code == 409


def test_the_database_itself_refuses_a_second_open_entry(
    client, admin_headers, tmp_path
):
    # Not only the service: the partial unique index is what makes the rule
    # hold for anything that writes to this database.
    project = make_project(client, admin_headers)
    check_in(client, admin_headers, project["id"])
    entry = client.get("/api/time/entries", headers=admin_headers).json()[0]

    db = sqlite3.connect(tmp_path / "test.db")
    try:
        db.execute(
            "INSERT INTO time_entries (user_id, project_id, started_at, ended_at,"
            " utc_offset) VALUES (1, ?, '2026-06-10 12:00:00', NULL, 0)",
            (entry["project_id"],),
        )
    except sqlite3.IntegrityError as error:
        message = str(error).lower()
        assert "uq_open_entry_per_project" in message or "unique" in message
    else:
        raise AssertionError("the database allowed two open entries on one project")
    finally:
        db.close()


def test_two_users_may_run_the_same_project_name(client, admin_headers):
    _, other = make_user(client, admin_headers, "colleague")
    mine = make_project(client, admin_headers, "Shared name")
    theirs = make_project(client, other, "Shared name")
    assert check_in(client, admin_headers, mine["id"]).status_code == 201
    assert check_in(client, other, theirs["id"]).status_code == 201


def test_a_duplicate_project_name_is_refused_for_one_user(client, admin_headers):
    make_project(client, admin_headers, "Twice")
    again = client.post("/api/projects", headers=admin_headers, json={"name": "Twice"})
    assert again.status_code == 409


def test_checking_out_when_nothing_runs_is_refused(client, admin_headers):
    project = make_project(client, admin_headers)
    assert check_out(client, admin_headers, project["id"]).status_code == 409


def test_a_session_ending_before_it_starts_is_refused(client, admin_headers):
    project = make_project(client, admin_headers)
    refused = record(client, admin_headers, project["id"], at(10, 17), at(10, 9))
    assert refused.status_code == 422


def test_a_session_over_midnight_is_split_across_both_days(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 22), at(11, 2))
    assert totals(client, admin_headers) == {
        ("2026-06-10", project["id"]): 2 * HOUR,
        ("2026-06-11", project["id"]): 2 * HOUR,
    }


def test_a_running_session_is_counted_up_to_as_of(client, admin_headers):
    project = make_project(client, admin_headers)
    check_in(client, admin_headers, project["id"], hour=9)
    counted = totals(client, admin_headers, as_of=at(10, 11))
    assert counted == {("2026-06-10", project["id"]): 2 * HOUR}


def test_parallel_sessions_are_added(client, admin_headers):
    work = make_project(client, admin_headers, "Work")
    meeting = make_project(client, admin_headers, "Meeting")
    record(client, admin_headers, work["id"], at(10, 9), at(10, 17))
    record(client, admin_headers, meeting["id"], at(10, 11), at(10, 12))
    # Nine tracked hours on a 24-hour day: that is what a sum over projects is.
    counted = totals(client, admin_headers)
    assert counted[("2026-06-10", work["id"])] == 8 * HOUR
    assert counted[("2026-06-10", meeting["id"])] == HOUR


def test_overlapping_sessions_on_one_project_are_refused(client, admin_headers):
    # Two projects at once is the point of the tracker; one project twice over
    # the same minutes reports the same hour twice under the same name.
    project = make_project(client, admin_headers)
    assert (
        record(client, admin_headers, project["id"], at(10, 9), at(10, 12)).status_code
        == 201
    )
    clash = record(client, admin_headers, project["id"], at(10, 11), at(10, 13))
    assert clash.status_code == 422
    assert "overlaps" in clash.json()["detail"]


def test_sessions_that_merely_touch_are_allowed(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))
    assert (
        record(client, admin_headers, project["id"], at(10, 12), at(10, 13)).status_code
        == 201
    )


def test_an_overlap_can_be_merged_instead(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))

    merged = client.post(
        "/api/time/entries",
        headers=admin_headers,
        json={
            "project_id": project["id"],
            "started_at": at(10, 11),
            "ended_at": at(10, 13),
            "utc_offset": 0,
            "merge_overlapping": True,
        },
    )
    assert merged.status_code == 201
    # One session covering both, earliest start to latest end - not two rows
    # counting 10:00 to 11:00 twice.
    rows = client.get("/api/time/entries", headers=admin_headers).json()
    assert len(rows) == 1
    assert rows[0]["started_at"].endswith("09:00:00")
    assert rows[0]["ended_at"].endswith("13:00:00")
    assert totals(client, admin_headers) == {("2026-06-10", project["id"]): 4 * HOUR}


def test_an_edit_into_an_overlap_is_refused_then_mergeable(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 11))
    record(client, admin_headers, project["id"], at(10, 14), at(10, 16))
    later = client.get("/api/time/entries", headers=admin_headers).json()[1]

    stretched = {"started_at": at(10, 10)}
    assert (
        client.put(
            f"/api/time/entries/{later['id']}", headers=admin_headers, json=stretched
        ).status_code
        == 422
    )

    assert (
        client.put(
            f"/api/time/entries/{later['id']}",
            headers=admin_headers,
            json={**stretched, "merge_overlapping": True},
        ).status_code
        == 200
    )
    rows = client.get("/api/time/entries", headers=admin_headers).json()
    assert len(rows) == 1
    assert rows[0]["started_at"].endswith("09:00:00")
    assert rows[0]["ended_at"].endswith("16:00:00")


def test_a_tag_totals_its_projects(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    one = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    two = make_project(client, admin_headers, "Reviews", tag_ids=[tag["id"]])
    record(client, admin_headers, one["id"], at(10, 9), at(10, 12))
    record(client, admin_headers, two["id"], at(10, 13), at(10, 14))
    assert totals(client, admin_headers, by="tag") == {
        ("2026-06-10", tag["id"]): 4 * HOUR
    }


def test_a_project_with_two_tags_counts_in_both(client, admin_headers):
    work = make_tag(client, admin_headers, "Work")
    meetings = make_tag(client, admin_headers, "Meetings")
    standup = make_project(
        client, admin_headers, "Standup", tag_ids=[work["id"], meetings["id"]]
    )
    record(client, admin_headers, standup["id"], at(10, 9), at(10, 10))
    assert totals(client, admin_headers, by="tag") == {
        ("2026-06-10", work["id"]): HOUR,
        ("2026-06-10", meetings["id"]): HOUR,
    }


def test_untagged_work_is_reported_under_no_tag(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    tagged = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    loose = make_project(client, admin_headers, "Reading")
    record(client, admin_headers, tagged["id"], at(10, 9), at(10, 10))
    record(client, admin_headers, loose["id"], at(10, 11), at(10, 13))
    assert totals(client, admin_headers, by="tag") == {
        ("2026-06-10", tag["id"]): HOUR,
        ("2026-06-10", None): 2 * HOUR,
    }


def test_deleting_a_tag_keeps_the_time(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    project = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))

    assert (
        client.delete(f"/api/tags/{tag['id']}", headers=admin_headers).status_code
        == 204
    )
    assert totals(client, admin_headers) == {("2026-06-10", project["id"]): 3 * HOUR}
    # The same time, now reported as untagged.
    assert totals(client, admin_headers, by="tag") == {("2026-06-10", None): 3 * HOUR}


def test_retagging_regroups_history(client, admin_headers):
    old = make_tag(client, admin_headers, "Old")
    new = make_tag(client, admin_headers, "New")
    project = make_project(client, admin_headers, "Backend", tag_ids=[old["id"]])
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))

    client.put(
        f"/api/projects/{project['id']}",
        headers=admin_headers,
        json={"tag_ids": [new["id"]]},
    )
    # A label is a view of the time, not a second record of it, so yesterday
    # regroups too.
    assert totals(client, admin_headers, by="tag") == {
        ("2026-06-10", new["id"]): 3 * HOUR
    }


def test_the_range_includes_a_session_merely_overlapping_it(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(9, 22), at(11, 3))
    counted = totals(client, admin_headers, start="2026-06-10", end="2026-06-10")
    assert counted == {("2026-06-10", project["id"]): 24 * HOUR}


def test_another_users_project_tag_and_entry_are_missing(client, admin_headers):
    _, other = make_user(client, admin_headers, "stranger")
    project = make_project(client, admin_headers)
    tag = make_tag(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))
    entry = client.get("/api/time/entries", headers=admin_headers).json()[0]

    assert (
        client.put(
            f"/api/projects/{project['id']}", headers=other, json={"name": "mine now"}
        ).status_code
        == 404
    )
    assert (
        client.put(
            f"/api/tags/{tag['id']}", headers=other, json={"name": "mine now"}
        ).status_code
        == 404
    )
    assert (
        client.delete(f"/api/time/entries/{entry['id']}", headers=other).status_code
        == 404
    )
    assert client.get("/api/time/entries", headers=other).json() == []


def test_deleting_a_tracked_project_is_refused(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))
    assert (
        client.delete(
            f"/api/projects/{project['id']}", headers=admin_headers
        ).status_code
        == 409
    )


def test_an_untracked_project_can_be_deleted(client, admin_headers):
    project = make_project(client, admin_headers)
    assert (
        client.delete(
            f"/api/projects/{project['id']}", headers=admin_headers
        ).status_code
        == 204
    )


def test_archiving_a_running_project_is_refused(client, admin_headers):
    project = make_project(client, admin_headers)
    check_in(client, admin_headers, project["id"])
    refused = client.put(
        f"/api/projects/{project['id']}", headers=admin_headers, json={"active": False}
    )
    assert refused.status_code == 409

    check_out(client, admin_headers, project["id"])
    assert (
        client.put(
            f"/api/projects/{project['id']}",
            headers=admin_headers,
            json={"active": False},
        ).status_code
        == 200
    )


def test_an_archived_project_cannot_be_checked_into(client, admin_headers):
    project = make_project(client, admin_headers)
    client.put(
        f"/api/projects/{project['id']}", headers=admin_headers, json={"active": False}
    )
    assert check_in(client, admin_headers, project["id"]).status_code == 409


def test_correcting_a_session_moves_the_total(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))
    entry = client.get("/api/time/entries", headers=admin_headers).json()[0]

    client.put(
        f"/api/time/entries/{entry['id']}",
        headers=admin_headers,
        json={"ended_at": at(10, 17)},
    )
    assert totals(client, admin_headers) == {("2026-06-10", project["id"]): 8 * HOUR}


def test_deleting_a_session_removes_its_time(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))
    entry = client.get("/api/time/entries", headers=admin_headers).json()[0]
    assert (
        client.delete(
            f"/api/time/entries/{entry['id']}", headers=admin_headers
        ).status_code
        == 204
    )
    assert totals(client, admin_headers) == {}


def test_the_export_agrees_with_the_summary(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    project = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    record(client, admin_headers, project["id"], at(10, 22), at(11, 2))

    response = client.get("/api/time/export.xlsx", headers=admin_headers)
    assert response.status_code == 200
    book = load_workbook(BytesIO(response.content))
    assert book.sheetnames == ["Sessions", "By project", "By tag"]

    by_project = list(book["By project"].values)[1:]
    assert by_project == [
        ("2026-06-10", "Backend", 2.0),
        ("2026-06-11", "Backend", 2.0),
    ]
    by_tag = list(book["By tag"].values)[1:]
    assert by_tag == [("2026-06-10", "Work", 2.0), ("2026-06-11", "Work", 2.0)]


def test_a_concurrent_second_check_in_reads_as_a_refusal(client, admin_headers):
    # The service check can lose a race between two devices; the partial index
    # is what actually holds, and its complaint must not surface as a 500.
    project = make_project(client, admin_headers)
    check_in(client, admin_headers, project["id"])

    from routers import time as time_router

    # Stand in for the racing request: the guard sees nothing running, because
    # the other request has not committed yet.
    original = time_router.running_entry
    time_router.running_entry = lambda *args, **kwargs: None
    try:
        raced = check_in(client, admin_headers, project["id"], hour=10)
    finally:
        time_router.running_entry = original

    assert raced.status_code == 409


def test_resuming_reopens_the_last_session(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))

    resumed = client.post(
        f"/api/projects/{project['id']}/resume", headers=admin_headers
    )
    assert resumed.status_code == 200
    # The same row, still starting at 09:00: the gap since 12:00 is absorbed
    # rather than left as a hole beside a new session.
    assert resumed.json()["ended_at"] is None
    assert resumed.json()["started_at"].endswith("09:00:00")
    assert len(client.get("/api/time/entries", headers=admin_headers).json()) == 1

    counted = totals(client, admin_headers, as_of=at(10, 14))
    assert counted == {("2026-06-10", project["id"]): 5 * HOUR}


def test_resuming_takes_the_most_recent_session(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 10))
    record(client, admin_headers, project["id"], at(10, 13), at(10, 14))

    resumed = client.post(
        f"/api/projects/{project['id']}/resume", headers=admin_headers
    ).json()
    assert resumed["started_at"].endswith("13:00:00")


def test_resuming_a_running_project_is_refused(client, admin_headers):
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 10))
    check_in(client, admin_headers, project["id"], hour=11)
    assert (
        client.post(
            f"/api/projects/{project['id']}/resume", headers=admin_headers
        ).status_code
        == 409
    )


def test_resuming_a_project_with_no_history_is_refused(client, admin_headers):
    project = make_project(client, admin_headers)
    assert (
        client.post(
            f"/api/projects/{project['id']}/resume", headers=admin_headers
        ).status_code
        == 409
    )


def test_another_users_project_cannot_be_resumed(client, admin_headers):
    _, other = make_user(client, admin_headers, "onlooker")
    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 10))
    assert (
        client.post(f"/api/projects/{project['id']}/resume", headers=other).status_code
        == 404
    )


def test_a_running_session_can_be_deleted(client, admin_headers):
    # The accidental tap: there is no end time to correct, so removing the row
    # has to be possible.
    project = make_project(client, admin_headers)
    check_in(client, admin_headers, project["id"])
    entry = client.get("/api/time/entries", headers=admin_headers).json()[0]

    assert (
        client.delete(
            f"/api/time/entries/{entry['id']}", headers=admin_headers
        ).status_code
        == 204
    )
    assert client.get("/api/time/entries", headers=admin_headers).json() == []
    # And the project is startable again, so the index released with the row.
    assert check_in(client, admin_headers, project["id"], hour=11).status_code == 201


def test_the_export_reads_in_local_time(client, admin_headers):
    project = make_project(client, admin_headers)
    # Two hours east: 22:00 UTC was midnight where it was recorded.
    record(client, admin_headers, project["id"], at(10, 22), at(11, 1), offset=120)

    response = client.get("/api/time/export.xlsx", headers=admin_headers)
    header, row = list(load_workbook(BytesIO(response.content))["Sessions"].values)

    assert header[1:7] == (
        "Started",
        "Ended",
        "Day offset",
        "Recorded offset",
        "Started (UTC)",
        "Ended (UTC)",
    )
    # Local first, as every screen shows it; UTC kept so it stays unambiguous.
    # The day's clock and the session's own are both named, because after a
    # flight they are not the same thing.
    assert row[1] == "2026-06-11 00:00"
    assert row[2] == "2026-06-11 03:00"
    assert row[3] == "UTC+02:00"
    assert row[4] == "UTC+02:00"
    assert row[5] == "2026-06-10 22:00"


def test_an_archived_project_leaves_the_summary(client, admin_headers):
    kept = make_project(client, admin_headers, "Backend")
    retired = make_project(client, admin_headers, "Reading")
    record(client, admin_headers, kept["id"], at(10, 9), at(10, 12))
    record(client, admin_headers, retired["id"], at(10, 20), at(10, 21))

    assert len(totals(client, admin_headers)) == 2
    client.put(
        f"/api/projects/{retired['id']}", headers=admin_headers, json={"active": False}
    )

    # Gone from the reports...
    assert totals(client, admin_headers) == {("2026-06-10", kept["id"]): 3 * HOUR}
    # ...but the hours are still recorded, and still exported.
    rows = client.get("/api/time/entries", headers=admin_headers).json()
    assert len(rows) == 2
    book = load_workbook(
        BytesIO(client.get("/api/time/export.xlsx", headers=admin_headers).content)
    )
    assert any("Reading" in str(row) for row in book["Sessions"].values)


def set_bands(client, headers, tag_id, bands):
    """Replace a tag's deduction rule."""
    return client.put(f"/api/tags/{tag_id}/deductions", headers=headers, json=bands)


def tag_rows(client, headers, **params):
    """Read the tag summary as ``{(day, tag): row}``."""
    response = client.get(
        "/api/time/summary", headers=headers, params={"by": "tag", **params}
    )
    return {(row["day"], row["key"]): row for row in response.json()}


def test_a_tag_with_no_rule_reports_what_it_tracked(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    project = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    record(client, admin_headers, project["id"], at(10, 9), at(10, 17))

    row = tag_rows(client, admin_headers)[("2026-06-10", tag["id"])]
    assert row["seconds"] == 8 * HOUR
    assert row["deduction"] == 0
    assert row["reported"] == 8 * HOUR


def test_a_capping_band_holds_the_day_at_its_threshold(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    project = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    assert (
        set_bands(
            client,
            admin_headers,
            tag["id"],
            [{"from_minutes": 600, "deduct_minutes": None}],
        ).status_code
        == 200
    )

    record(client, admin_headers, project["id"], at(10, 8), at(10, 22))
    row = tag_rows(client, admin_headers)[("2026-06-10", tag["id"])]
    assert (row["seconds"], row["deduction"], row["reported"]) == (
        14 * HOUR,
        4 * HOUR,
        10 * HOUR,
    )

    stored = client.get(
        f"/api/tags/{tag['id']}/deductions", headers=admin_headers
    ).json()
    assert stored == [{"from_minutes": 600, "deduct_minutes": None}]


def test_a_lunch_rule_deducts_by_band(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    project = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    assert (
        set_bands(
            client,
            admin_headers,
            tag["id"],
            [
                {"from_minutes": 0, "deduct_minutes": 30},
                {"from_minutes": 360, "deduct_minutes": 45},
            ],
        ).status_code
        == 200
    )

    # A short day: half an hour.
    record(client, admin_headers, project["id"], at(10, 9), at(10, 13))
    short = tag_rows(client, admin_headers)[("2026-06-10", tag["id"])]
    assert (short["seconds"], short["deduction"], short["reported"]) == (
        4 * HOUR,
        1800,
        4 * HOUR - 1800,
    )

    # A long one: three quarters.
    record(client, admin_headers, project["id"], at(11, 9), at(11, 18))
    long_day = tag_rows(client, admin_headers)[("2026-06-11", tag["id"])]
    assert (long_day["seconds"], long_day["deduction"], long_day["reported"]) == (
        9 * HOUR,
        2700,
        9 * HOUR - 2700,
    )


def test_the_rule_is_retroactive(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    project = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    record(client, admin_headers, project["id"], at(10, 9), at(10, 17))

    set_bands(
        client, admin_headers, tag["id"], [{"from_minutes": 0, "deduct_minutes": 60}]
    )
    # Nothing was recomputed, because nothing was stored.
    assert tag_rows(client, admin_headers)[("2026-06-10", tag["id"])]["reported"] == (
        7 * HOUR
    )


def test_the_rule_belongs_to_its_tag_alone(client, admin_headers):
    work = make_tag(client, admin_headers, "Work")
    reading = make_tag(client, admin_headers, "Reading")
    worked = make_project(client, admin_headers, "Backend", tag_ids=[work["id"]])
    read = make_project(client, admin_headers, "A book", tag_ids=[reading["id"]])
    set_bands(
        client, admin_headers, work["id"], [{"from_minutes": 0, "deduct_minutes": 45}]
    )

    record(client, admin_headers, worked["id"], at(10, 9), at(10, 17))
    record(client, admin_headers, read["id"], at(10, 19), at(10, 21))

    rows = tag_rows(client, admin_headers)
    assert rows[("2026-06-10", work["id"])]["reported"] == 8 * HOUR - 2700
    # A day of reading owes nobody a lunch break.
    assert rows[("2026-06-10", reading["id"])]["reported"] == 2 * HOUR


def test_project_rows_carry_no_deduction(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    project = make_project(client, admin_headers, "Backend", tag_ids=[tag["id"]])
    set_bands(
        client, admin_headers, tag["id"], [{"from_minutes": 0, "deduct_minutes": 45}]
    )
    record(client, admin_headers, project["id"], at(10, 9), at(10, 17))

    row = next(
        row
        for row in client.get("/api/time/summary", headers=admin_headers).json()
        if row["key"] == project["id"]
    )
    # A deduction belongs to a tag; a project bar stays tracked time.
    assert (row["deduction"], row["reported"]) == (0, 8 * HOUR)


def test_two_bands_at_the_same_threshold_are_refused(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    refused = set_bands(
        client,
        admin_headers,
        tag["id"],
        [
            {"from_minutes": 0, "deduct_minutes": 30},
            {"from_minutes": 0, "deduct_minutes": 45},
        ],
    )
    assert refused.status_code == 422


def test_deleting_a_tag_takes_its_rule(client, admin_headers):
    tag = make_tag(client, admin_headers, "Work")
    set_bands(
        client, admin_headers, tag["id"], [{"from_minutes": 0, "deduct_minutes": 30}]
    )
    assert (
        client.delete(f"/api/tags/{tag['id']}", headers=admin_headers).status_code
        == 204
    )


def test_another_users_rule_is_missing(client, admin_headers):
    _, other = make_user(client, admin_headers, "bystander")
    tag = make_tag(client, admin_headers, "Work")
    assert (
        client.get(f"/api/tags/{tag['id']}/deductions", headers=other).status_code
        == 404
    )
    assert set_bands(client, other, tag["id"], []).status_code == 404


def test_the_tracked_range_reports_the_edges(client, admin_headers):
    empty = client.get("/api/time/range", headers=admin_headers).json()
    assert empty == {"first": None, "last": None}

    project = make_project(client, admin_headers)
    record(client, admin_headers, project["id"], at(10, 9), at(10, 12))
    record(client, admin_headers, project["id"], at(14, 9), at(14, 12))

    assert client.get("/api/time/range", headers=admin_headers).json() == {
        "first": "2026-06-10",
        "last": "2026-06-14",
    }


def test_the_tracked_range_reads_each_days_own_offset(client, admin_headers):
    project = make_project(client, admin_headers)
    # 23:30 UTC is already the next day two hours east.
    record(client, admin_headers, project["id"], at(10, 23, 30), at(11, 1), offset=120)
    assert client.get("/api/time/range", headers=admin_headers).json()["first"] == (
        "2026-06-11"
    )
